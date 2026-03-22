import csv, os, glob

sage = r"D:\raw_data\Sage_Data"
for f in os.listdir(sage):
    fp = os.path.join(sage, f)
    if f.endswith(".csv"):
        with open(fp, encoding="utf-8", errors="replace") as fh:
            r = csv.reader(fh)
            hdr = next(r, None)
            row1 = next(r, None)
            print(f"=== {f} ===")
            print(f"  HEADERS: {hdr}")
            print(f"  ROW1: {row1}")
            print()

# CPI xlsx via openpyxl
try:
    import openpyxl
    for xf in glob.glob(r"D:\raw_data\cpi\*.xlsx"):
        wb = openpyxl.load_workbook(xf, read_only=True)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(list(row))
            if i >= 4: break
        print(f"=== {os.path.basename(xf)} ===")
        for r in rows:
            print(f"  {r}")
        print()
        wb.close()
except Exception as e:
    print(f"openpyxl error: {e}")

# Religion census xlsx
try:
    for xf in glob.glob(os.path.join(sage, "*.xlsx")):
        wb = openpyxl.load_workbook(xf, read_only=True)
        for sn in wb.sheetnames[:2]:
            ws = wb[sn]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(list(row))
                if i >= 2: break
            print(f"=== {os.path.basename(xf)} | {sn} ===")
            for r in rows:
                print(f"  {r}")
            print()
        wb.close()
except Exception as e:
    print(f"religion xlsx error: {e}")

# HSUS
hsus = r"D:\raw_data\HSUS\A-Population"
if os.path.isdir(hsus):
    for f in os.listdir(hsus)[:5]:
        fp = os.path.join(hsus, f)
        if f.endswith(".csv"):
            with open(fp, encoding="utf-8", errors="replace") as fh:
                r = csv.reader(fh)
                hdr = next(r, None)
                print(f"=== HSUS/{f} ===")
                print(f"  {hdr}")
